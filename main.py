import PyPDF2

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers, PatternFill, Border, Side
from pathlib import Path 
from PyPDF2 import PdfFileReader
from tkinter import filedialog, Tk    

chequingAccountNumber = 54321
savingAccountNumber = 12345
companyName = 'Placeholder Company Name'

#region  Excel Setup
wb = Workbook()
wsCheque = wb['Sheet']
wsCheque.title = "reconciliation-Cheque"
wsSaving = wb.create_sheet("reconciliation-Saving AC")

for sheet in wb:
    for row in range(1, 50):
        for column in range(len('ABCDEFG')):
            sheet[f'{"ABCDEFG"[column]}{row}'].font = Font(name = 'Times New Roman', size = 12)

    for row in range(6, 50):
        for column in range(len('DEF')):
            sheet[f'{"DEF"[column]}{row}'].alignment = Alignment(horizontal = 'right')
        
    sheet.merge_cells('B1:G1')
    sheet['B1'].alignment = Alignment(horizontal = 'center')

    sheet.merge_cells('B2:G2')
    sheet['B2'].alignment = Alignment(horizontal = 'center')

    sheet.merge_cells('B3:G3')
    sheet['B3'].alignment = Alignment(horizontal = 'center')

    sheet.merge_cells('A5:C5')

    sheet.column_dimensions['A'].width = 4.13
    sheet.column_dimensions['B'].width = 12.38
    sheet.column_dimensions['C'].width = 35.5
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12

    for column in range(len('ABCDEFG')):
        sheet[f'{"ABCDEFG"[column]}{5}'].border = Border(top = Side(border_style='thin'))
    

    sheet.row_dimensions[6].height = 22.5
    for column in range(len('ABCDEFG')):
        sheet[f'{"ABCDEFG"[column]}{6}'].fill = PatternFill("solid", start_color="C0C0C0")
        
    sheet['B1'] = f"{companyName.capitalize}"

    if sheet == wsCheque:
        sheet['B3'] = f"CIBC Canada Client/Trust Account-Cheque-{chequingAccountNumber}"
    else:
        sheet['B3'] = f"CIBC Canada Client/Trust Account-Saving Account-{savingAccountNumber}"

    sheet['A5'].font = Font(name = 'Times New Roman', size = 12, bold = True)
    sheet['A6'] = "No." ; sheet['B6'] = "Date"; sheet['C6'] = "Description"; sheet['D6'] = "Withdrawal"; sheet['E6'] = "Deposits"; sheet['F6'] = "Balance"; sheet['G6'] = "Assigned To"; 
        
    
#endregion

for i in range(2):  

    accountType = input('Chequing or savings: ')
    print(f'Please select {accountType} file')
    if accountType.lower == "chequing":
        wsActive = wsCheque
    else: 
        wsActive = wsSaving


    # print(filename)
    # print(type(filename))
    # Tk().withdraw() 
    filename = filedialog.askopenfilename()

    try:
        pdf = PdfFileReader(filename)

    except PyPDF2.errors.PdfReadError: 
        print("Error reading PDF")
        
    else:
        previousBalance = 0

        #region Converting PDF to text file 
        with Path('lorem_text.txt').open(mode ='w') as initialFile:
            initialFile.truncate(0)
            text = ''
            for page in pdf.pages:
                text += page.extractText()
            initialFile.write(text)
        #endregion 
            
        #region Obtaining closing balance and date 
        with Path('lorem_text.txt').open(mode ='r') as initialFile:
            initialFileRead = initialFile.readlines()

            closingBalanceDate = (((initialFileRead[34][19: 50])).strip()) 
            month = closingBalanceDate[0:3]

            closingBalance = (initialFileRead[36]) 
            sheet['B2'] = f"Bank Reconcilation, {closingBalanceDate}"
            sheet['A5'] = f"Cleared Transactions Per Bank Statement {closingBalanceDate}"
        #endregion

        #region Creating condensed copy of original text file
        with Path('lorem_text.txt').open(mode ='r') as initialFile:    
            with Path('lorem_text_copy.txt').open(mode ='w') as copiedFile:
                copy = False
                nextLineIB = False 
                skip = 0
                for line in initialFile:
                    #Recording opening balance 
                    if nextLineIB:
                        previousBalance = line[1:]
                        wsActive['B7'] = f'{month} 1'
                        wsActive['C7'] = 'Opening Balance'
                        wsActive['F7'] = previousBalance
                        nextLineIB = False 

                    if skip != 0:
                        skip -= 1
                        continue             

                    if line.strip() == "Opening balance":
                        skip = 1
                        nextLineIB = True 
                        copy = True
                    
                    if line.strip() == "Closing balance":
                        copy = False

                    if line.strip() == "(continued on next page)":
                        copy = False

                    if line.strip() == "Balance forward":
                        skip = 1   
                        copy = True

                    if copy and skip == 0:
                        copiedFile.write(line)
        #endregion

        #region Adding in transaction dates for all transactions
        with Path('lorem_text_copy.txt').open(mode ='r') as copiedFile: 
                copiedFileRead = copiedFile.readlines()
                currentDate = ""
                
                for line in copiedFileRead:
                    index = copiedFileRead.index(line)          
                    if month in line:
                        currentDate = line
                    if line.isspace():
                        copiedFileRead[index] = currentDate
        #endregion

        with Path('lorem_text_copy.txt').open(mode ='w') as copiedFile:
            copiedFile.writelines(copiedFileRead)

        #region Splitting transactions into groups and moving data into Excel 
        with Path('lorem_text_copy.txt').open(mode ='r') as copiedFile: 
                
                copiedFileRead = copiedFile.readlines()

                newGroup = True
                group = []
                totalGroup = 0

                def FinalizeGroup(group):
                    if len(group) > 0:
                        print(group)
                    print("Balance: " + str(previousBalance))
                    group.clear()

                for line in copiedFileRead:
                    if line.startswith(month):
                        if len(group) > 0: 
                            wsActive[f'B{7 + totalGroup}'] = group[0]

                            wsActive[f'C{7 + totalGroup}'] = group[1]

                            if float(group[-1].replace(',','')) < float(previousBalance.replace(',','')):
                                wsActive[f'D{7 + totalGroup}'] = float(group[-2].replace(',',''))
                                wsActive[f'D{7 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00
                            else: 
                                wsActive[f'E{7 + totalGroup}'] = float(group[-2].replace(',',''))
                                wsActive[f'E{7 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00

                            wsActive[f'F{7 + totalGroup}'] = float(group[-1].replace(',',''))
                            wsActive[f'F{7 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00
                            
                            previousBalance = group[-1]

                        wsActive[f'A{7 + totalGroup}'] = totalGroup + 1
                        wsActive[f'A{7 + totalGroup}'].alignment = Alignment(horizontal='center')

                        FinalizeGroup(group)
                        totalGroup += 1

                    group.append(line[0: len(line) - 1])

                print(group) #Printing out last group with no month behind it 
                wsActive[f'A{7 + totalGroup}'] = totalGroup + 1
                wsActive[f'A{7 + totalGroup}'].alignment = Alignment(horizontal='center')

                wsActive[f'B{7 + totalGroup}'] = group[0]
                
                wsActive[f'C{7 + totalGroup}'] = group[1]

                if float(group[-1].replace(',','')) < float(previousBalance.replace(',','')):
                    wsActive[f'D{7 + totalGroup}'] = float(group[-2].replace(',',''))
                    wsActive[f'D{7 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00
                else: 
                    wsActive[f'E{7 + totalGroup}'] = float(group[-2].replace(',',''))
                    wsActive[f'E{7 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00

                wsActive[f'F{7 + totalGroup}'] = float(group[-1].replace(',',''))
                wsActive[f'F{7 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00

                print('Total number of groups: ' + str(totalGroup))
        #endregion

        #region Formatting end of spreadsheet
        print(f'Total group is ==== {totalGroup}')
        wsActive[f'B{9 + totalGroup}'] = f'Total withdrawals made and cleared as of {closingBalanceDate}'
        wsActive[f'D{9 + totalGroup}'] = f'=sum(D7:D{7 + totalGroup})'
        wsActive[f'D{9 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00

        wsActive[f'B{10 + totalGroup}'] = f'Total deposits made and cleared as of {closingBalanceDate}'
        wsActive[f'E{10 + totalGroup}'] = f'=sum(E7:E{7 + totalGroup})'
        wsActive[f'E{10 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00

        wsActive[f'B{11 + totalGroup}'] = f'Balance of cleared transactions as of {closingBalanceDate}'
        wsActive[f'F{11 + totalGroup}'] = group[-1]
        wsActive[f'F{11 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00

        wsActive[f'A{13 + totalGroup}'] = f'Uncleared transactions {closingBalanceDate}'
        wsActive[f'A{13 + totalGroup}'].font = Font(name = 'Times New Roman', size = 12, bold = True)
        for column in range(len('ABCDFEG')):
            wsActive[f'{"ABCDEFG"[column]}{13 + totalGroup}'].border = Border(top = Side(border_style='thin'))

        wsActive[f'A{14 + totalGroup}'] = "No." ; wsActive[f'B{14 + totalGroup}'] = "Date"; wsActive[f'C{14 + totalGroup}'] = "Description"; wsActive[f'D{14 + totalGroup}'] = "Withdrawal"; wsActive[f'E{14 + totalGroup}'] = "Deposits"; wsActive[f'F{14 + totalGroup}'] = "Balance"; wsActive[f'G{14 + totalGroup}'] = "Assigned To";
        wsActive.row_dimensions[14 + totalGroup].height = 22.5
        for column in range(len('ABCDEFG')):
            wsActive[f'{"ABCDEFG"[column]}{14 + totalGroup}'].fill = PatternFill("solid", start_color="C0C0C0")

        wsActive[f'B{18 + totalGroup}'] = f'Total withdrawals made but not cleared as of {closingBalanceDate}'
        wsActive[f'D{18 + totalGroup}'] = 0

        wsActive[f'B{19 + totalGroup}'] = f'Total deposits made but not cleared as of {closingBalanceDate}'
        wsActive[f'E{19 + totalGroup}'] = 0

        wsActive[f'B{20 + totalGroup}'] = f'Balance of uncleared transactions as of {closingBalanceDate}'
        wsActive[f'F{20 + totalGroup}'] = 0

        wsActive[f'B{22 + totalGroup}'] = f'True cash balance as of {closingBalanceDate}'
        wsActive[f'F{22 + totalGroup}'] = group[-1]
        wsActive[f'F{22 + totalGroup}'].number_format = numbers.FORMAT_NUMBER_00

        for row in range(5, 22 + totalGroup):
            wsActive[f'G{row}'].border = Border(right= Side(border_style='thin'))

        for column in range(len('ABCDFEG')):
            wsActive[f'{"ABCDEFG"[column]}{22 + totalGroup}'].border = Border(bottom = Side(border_style='thin'))

        #Adding in corner borders
        wsActive[f'G5'].border = Border(right = Side(border_style='thin'), top = Side(border_style='thin'))
        wsActive[f'G{13 + totalGroup}'].border = Border(right = Side(border_style='thin'), top = Side(border_style='thin'))
        wsActive[f'G{22 + totalGroup}'].border = Border(right = Side(border_style='thin'), bottom = Side(border_style='thin'))

        #endregion

wb.save('lorem.xlsx')
    

